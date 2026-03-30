import pandas as pd
import numpy as np
import seaborn as sb
import matplotlib.pyplot as plt
import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment  
from openpyxl.styles import PatternFill, Font
import xlwings as xw
from datetime import datetime

############################
### Function Definitions ###
############################
def autosize_df_columns(ws, df, start_col=1):
    for i, col in enumerate(df.columns, start=start_col):
        max_length = len(str(col))
        for val in df[col]:
            if val is not None:
                max_length = max(max_length, len(str(val)))
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

# Excel Generation function
def generate_excel(destination, excel_sheet_name, df, header_row):
    book = load_workbook(destination, keep_vba=True)
    with pd.ExcelWriter(destination, engine='openpyxl', mode='a', if_sheet_exists='overlay', engine_kwargs={"keep_vba": True}) as writer:
        sheet = excel_sheet_name
        df.to_excel(writer, sheet_name=sheet, index=False)
        ws = writer.sheets[sheet]

        header_row = header_row
        first_data_row = header_row + 1
        last_data_row = ws.max_row
        start_col = 1
        last_col = df.shape[1]

        ws.auto_filter.ref = f"{get_column_letter(start_col)}{header_row}:{get_column_letter(last_col)}{last_data_row}"

        wb = writer.book
        ws = wb[sheet]
        autosize_df_columns(ws, df)

# delete sheet in excel sheet function
def delete_excel_sheet(destination, excel_sheet_name):
    wb = load_workbook(destination, keep_vba=True)
    if excel_sheet_name in wb.sheetnames:
        del wb[excel_sheet_name]
    wb.save(destination)

# Apply color to the different groups for differentiation
def apply_group_colors(destination, sheet_name):
    wb = load_workbook(destination, keep_vba=True)
    ws = wb[sheet_name]
    colors = [
        PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
        PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
        PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ]
    last_group = None
    color_idx = -1
    for row_num in range(2, ws.max_row + 1):
        current_group = ws.cell(row=row_num, column=4).value
        
        if current_group != last_group:
            color_idx = (color_idx + 1) % len(colors)
            
        for col_num in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col_num).fill = colors[color_idx]
            
        last_group = current_group
    wb.save(destination)

# Excel Clickable Link function (for clicking and changing to the relevant sheet)
def excel_clickable_cell(val, sheet, cell="A1"):
    if val is None or str(val).strip() == "":
        return val
    return f'=HYPERLINK("#\'{sheet}\'!{cell}", "{val}")'

# Chart functions
def merge_cells_title(sheet, cell1, cell2, row_no, col, val, horizontal_val, vertical_val, color):
    sheet.merge_cells(f'{cell1}:{cell2}')
    cell = sheet.cell(row=row_no, column=col)
    cell.value = val
    cell.alignment = Alignment(horizontal=horizontal_val, vertical=vertical_val)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

def hyperlink_cell(ws, first_data_row, last_data_row, column_no=1):
    for row in range(first_data_row, last_data_row + 1):
        ws.cell(row=row, column=column_no).style = "Hyperlink"

def univariate_table(df, variable, top_n=5):
    counts = df[variable].value_counts().reset_index()
    counts.columns = [variable, 'count']
    if len(counts) > top_n:
        sub_df = counts.head(top_n).copy()
        misc_count = counts.iloc[top_n:]['count'].sum()
        misc_row = pd.DataFrame({variable: ['Misc'], 'count': [misc_count]})
        sub_df = pd.concat([sub_df, misc_row], ignore_index=True)
    else:
        sub_df = counts
    sub_df_rows = sub_df.shape[0]
    sub_df_columns = sub_df.shape[1]
    return sub_df, sub_df_rows, sub_df_columns

def pivot_table(df, index_col, col, val):
    pivot_table_df = pd.pivot_table(df, index=index_col, columns=col, values=val, aggfunc="count", fill_value=0).stack().reset_index()
    if isinstance(index_col, list):
        new_cols = index_col + [col, "count"]
    else:
        new_cols = [index_col, col, "count"]
    pivot_table_df.columns = new_cols
    return pivot_table_df

def pivot_table_wide(df, index_col, col, val, sheet=None, cell=None):
    column_order = ["Critical", "High", "Medium", "Low"]
    pivot_df = pivot_table(df, index_col, col, val)
    pivot_df_wide = (
        pivot_df
        .pivot(index=index_col, columns=col, values="count")
        .fillna(0)
    )
    existing_cols = [c for c in column_order if c in pivot_df_wide.columns]
    pivot_df_wide = pivot_df_wide.reindex(columns=existing_cols)
    pivot_df_wide = pivot_df_wide.reset_index()
    if isinstance(index_col, list):
        pivot_df_wide.insert(2, 'Label',
                             pivot_df_wide[index_col[0]].astype(str).str.strip()
                             + " - " +
                             pivot_df_wide[index_col[1]].astype(str).str.strip()
                             )
    pivot_df_wide_rows = pivot_df_wide.shape[0]
    pivot_df_wide_columns = pivot_df_wide.shape[1]
    return pivot_df_wide, pivot_df_wide_rows, pivot_df_wide_columns

def barchart_creation(sheet, chartType, variable, x_title, y_title, min_col, max_col, min_row, max_row, showVal, showSerName, showCatName, showLeaderLines, cell, chartStyle=None, chartGrouping=None, chartOverlap=None, shape=None):
    chart = BarChart()
    chart.type = chartType
    if chartStyle is not None:
        chart.style = chartStyle
    if chartGrouping is not None:
        chart.grouping = chartGrouping
    if chartOverlap is not None:
        chart.overlap = chartOverlap
    chart.title = f"Vulnerabilities by {variable}"
    num_categories = max_row - min_row

    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.y_axis.majorGridlines = None
    chart.x_axis.title.overlay = chart.y_axis.title.overlay = chart.legend.overlay = chart.title.overlay = False
    
    # Data & Categories
    data = Reference(sheet, min_col=min_col, min_row=min_row, max_row=max_row, max_col=max_col)
    cats = Reference(sheet, min_col=min_col-1, min_row=min_row+1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.dataLabels = DataLabelList() 
    chart.dataLabels.showVal = showVal
    chart.dataLabels.showSerName = showSerName
    chart.dataLabels.showCatName = showCatName
    chart.dataLabels.showLeaderLines = showLeaderLines
    chart.set_categories(cats)
    
    if chartGrouping is not None:
        chart.x_axis.delete = False
        chart.x_axis.tickLblPos = "nextTo" # position of the x axis labels 
        chart.x_axis.crosses = "autoZero" # position of the 0 point where x axis and y axis cross
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickMarkSkip = 1
    else:
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.crosses = "min"
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickMarkSkip = 1
    
    chart.width = 6 + (num_categories * 2) 
    chart.height = 7 + (num_categories * 0.1)
    
    sheet.add_chart(chart, cell)
    return chart

def piechart_creation(sheet, shape, variable, min_col, min_row, max_row, max_col, showVal, showSerName, showCatName, showPercent, cell):
    chart = PieChart()
    chart.title = f"Vulnerabilities by {variable}"
    num_categories = max_row - min_row

    # Data & Categories
    data = Reference(sheet, min_col=min_col, min_row=min_row, max_row=max_row, max_col=max_col)
    cats = Reference(sheet, min_col=min_col-1, min_row=min_row+1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.dataLabels = DataLabelList() 
    chart.dataLabels.showVal = showVal
    chart.dataLabels.showSerName = showSerName
    chart.dataLabels.showCatName = showCatName
    chart.dataLabels.showPercent = showPercent
    chart.set_categories(cats)

    chart.shape = shape
    chart.width = 6 + (num_categories * 2) 
    chart.height = 7 + (num_categories * 0.1)
    sheet.add_chart(chart, cell)
    return chart
############################

today = datetime.today()
new_file_name = "dummy_data"
raw_df = pd.read_excel(f"{new_file_name}_raw.xlsx")
df = raw_df.copy()

# insert 'scan' column defined based on the value in 'port'
target_position = df.columns.get_loc('port') + 1
target_values = np.where(df['port'] != 0, 'Network Scan', 'Agent Scan')
df.insert(loc=target_position, column='scan', value=target_values)

newfile = f"{new_file_name}.xlsm"
if os.path.exists(newfile):
    os.remove(newfile)
source_file = "./template.xlsm"
shutil.copy(source_file, newfile)
severity_thresholds = {"Critical": 60, "High": 60, "Medium": 90, "Low": 90}

# time columns conversion
df['patch_publication_date'] = pd.to_datetime(df['patch_publication_date'], format = "%d/%m/%Y %H:%M:%S")
df['first_observed_date'] = pd.to_datetime(df['first_observed_date'], format="%d/%m/%Y %H:%M:%S")

# create a new column for today's date
df['comparison_date'] = pd.to_datetime(today, format="%d/%m/%Y %H:%M:%S")

# create a new column to calculate the difference in days
for idx, row in df.iterrows():
    if pd.isnull(row['patch_publication_date']):
        df.at[idx, "difference"] = (row['comparison_date'] - row['first_observed_date']).days
    else:
        df.at[idx, "difference"] = (row['comparison_date'] - row['patch_publication_date']).days

# create a new column to apply the overdue flah (Y or N)
df["overdue"] = df.apply(
    lambda row: "Y" if row["difference"] > severity_thresholds[row["severity"]] else "N",
    axis=1
)

overdue_df = df.loc[df['overdue'] == 'Y'].reset_index(drop=True)
non_overdue_df = df.loc[df['overdue'] == 'N'].reset_index(drop=True)

# create a column to calculate the number of days overdue/days to overdue
overdue_df["Days Overdued By (Days)"] = overdue_df['difference'] - overdue_df['severity'].map(severity_thresholds)
non_overdue_df["Days to Overdue (Days)"] = non_overdue_df['severity'].map(severity_thresholds) - non_overdue_df['difference']

raw_df = raw_df.sort_values(by=["asset_group", "plugin_family"]).reset_index(drop=True)
asset_group_rows = raw_df.index[raw_df["asset_group"].ne(raw_df["asset_group"].shift())]
# print(asset_group_rows)
asset_group_family_rows = raw_df.index[raw_df[["asset_group", "plugin_family"]].ne(raw_df[["asset_group", "plugin_family"]].shift()).any(axis=1)]
# print(asset_group_family_rows)
asset_group_dict = {raw_df.loc[idx, "asset_group"]: idx+2 for idx in asset_group_rows} # retrieve the rows that the new asset_group start at in the table for hyperlink later on, +2 because of pandas indexing and header row
# print(asset_group_dict)
asset_group_family_dict = {f"{raw_df.loc[idx, 'asset_group']} - {raw_df.loc[idx, 'plugin_family']}": idx+2 for idx in asset_group_family_rows}
# print(asset_group_family_dict)

# place this new table with additional columns into a new sheet
sheet_name = "Overdue Vulnerabilities"
generate_excel(newfile, sheet_name, overdue_df, 1)
sheet_name = "Non-Overdue Vulnerabilities"
generate_excel(newfile, sheet_name, non_overdue_df, 1)
sheet_name = "Original Vulnerabilities"
generate_excel(newfile, sheet_name, raw_df, 1)
apply_group_colors(newfile, sheet_name)
delete_excel_sheet(newfile, "Sheet1")

def main():
    family_vul, family_vul_rows, family_vul_columns = univariate_table(df, 'plugin_family')
    severity_vul, severity_vul_rows, severity_vul_columns = univariate_table(df, 'severity')
    asset_group_vul, asset_group_vul_rows, asset_group_vul_columns = univariate_table(df, 'asset_group')
    asset_group_vul['asset_group'] = asset_group_vul['asset_group'].apply(lambda x: excel_clickable_cell(x, sheet="Original Vulnerabilities", cell=f"A{asset_group_dict.get(x, 1)}"))
    overdue_vul, overdue_vul_rows, overdue_vul_columns = univariate_table(df, 'overdue')
    overdue_vul['overdue'] = np.where(overdue_vul['overdue'] == "Y", overdue_vul['overdue'].apply(excel_clickable_cell, sheet="Overdue Vulnerabilities"), overdue_vul['overdue'].apply(excel_clickable_cell, sheet="Non-Overdue Vulnerabilities"))
    barchart_start_column_value = family_vul_columns+2
    barchart_start_column = get_column_letter(barchart_start_column_value)
    pivot_table_1_wide, pivot_table_1_wide_rows, pivot_table_1_wide_columns = pivot_table_wide(df, "plugin_family", "severity", "plugin_id")
    pivot_table_2_wide, pivot_table_2_wide_rows, pivot_table_2_wide_columns = pivot_table_wide(df, "asset_group", "severity", "plugin_id")
    pivot_table_2_wide['asset_group'] = pivot_table_2_wide['asset_group'].apply(lambda x: excel_clickable_cell(x, sheet="Original Vulnerabilities", cell=f"A{asset_group_dict.get(x, 1)}"))
    pivot_table_3_wide, pivot_table_3_wide_rows, pivot_table_3_wide_columns = pivot_table_wide(df, "overdue", "severity", "plugin_id")
    pivot_table_3_wide['overdue'] = np.where(pivot_table_3_wide['overdue'] == "Y", pivot_table_3_wide['overdue'].apply(excel_clickable_cell, sheet="Overdue Vulnerabilities"), pivot_table_3_wide['overdue'].apply(excel_clickable_cell, sheet="Non-Overdue Vulnerabilities"))
    pivot_table_4_wide, pivot_table_4_wide_rows, pivot_table_4_wide_columns = pivot_table_wide(df, ["asset_group", "plugin_family"], "severity", "plugin_id")
    pivot_table_4_wide['Label'] = pivot_table_4_wide['Label'].apply(lambda x: excel_clickable_cell(x, sheet="Original Vulnerabilities", cell=f"A{asset_group_family_dict.get(x, 1)}"))

    sheet_name = "Summary"
    book = load_workbook(newfile, keep_vba=True)
    with pd.ExcelWriter(newfile, engine="openpyxl", mode="a", if_sheet_exists="overlay", engine_kwargs={"keep_vba": True}) as writer:
        sheet = sheet_name
        new_row = 3
        overdue_vul.to_excel(writer, sheet_name=sheet, startrow=new_row, startcol=0, index=False)
        ws = writer.sheets[sheet]
        header_row = new_row+1
        first_data_row = header_row+1
        last_data_row = header_row+overdue_vul_rows
        hyperlink_cell(ws, first_data_row, last_data_row)
        new_row = new_row + overdue_vul_rows + 2
        severity_vul.to_excel(writer, sheet_name=sheet, startrow=new_row, startcol=0, index=False)
        new_row = new_row + severity_vul_rows + 2
        asset_group_vul.to_excel(writer, sheet_name=sheet, startrow=new_row, startcol=0, index=False)
        header_row = new_row+1
        first_data_row = header_row+1
        last_data_row = header_row+asset_group_vul_rows
        hyperlink_cell(ws, first_data_row, last_data_row)
        new_row = new_row + asset_group_vul_rows + 2
        family_vul.to_excel(writer, sheet_name=sheet, startrow=new_row, startcol=0, index=False)

        wb = writer.book
        ws = wb[sheet]
        autosize_df_columns(ws, family_vul)
        wb.move_sheet(ws, offset=-wb.sheetnames.index(sheet))

    wb = load_workbook(newfile, keep_vba=True)
    sheet = wb[sheet_name]

    charts_length = []

    min_row_table = 4
    max_row_table = min_row_table + overdue_vul_rows
    overdue_vul_chart = barchart_creation(sheet, "col", "Overdue", "Overdue", "Count", 2, overdue_vul_columns, min_row_table, max_row_table, True, False, False, False, f"{barchart_start_column}4", chartStyle=10, shape=4)
    overdue_vul_piechart = piechart_creation(sheet, 4, "Overdue", 2, min_row_table, max_row_table, overdue_vul_columns, False, False, False, True, f"{get_column_letter(barchart_start_column_value + round(overdue_vul_chart.width / 1.7)+1)}4")
    charts_length.append(barchart_start_column_value + round(overdue_vul_chart.width / 1.7)+1+round(overdue_vul_piechart.width / 1.7))
    min_row_table = min_row_table + overdue_vul_rows + 2
    max_row_table = max_row_table + 2 + severity_vul_rows
    severity_vul_chart = barchart_creation(sheet, "col", "Severity", "Severity", "Count", 2, severity_vul_columns, min_row_table, max_row_table, True, False, False, False, f"{barchart_start_column}21", chartStyle=10, shape=4)
    severity_vul_piechart = piechart_creation(sheet, 4, "Severity", 2, min_row_table, max_row_table, severity_vul_columns, False, False, False, True, f"{get_column_letter(barchart_start_column_value + round(severity_vul_chart.width / 1.7)+1)}21")
    charts_length.append(barchart_start_column_value + round(severity_vul_chart.width / 1.7)+1+round(severity_vul_piechart.width / 1.7))
    min_row_table = min_row_table + severity_vul_rows + 2
    max_row_table = max_row_table + 2 + asset_group_vul_rows
    asset_group_vul_chart = barchart_creation(sheet, "col", "Asset Group", "Asset Group", "Count", 2, asset_group_vul_columns, min_row_table, max_row_table, True, False, False, False, f"{barchart_start_column}38", chartStyle=10, shape=4)
    asset_group_vul_piechart = piechart_creation(sheet, 4, "Asset Group", 2, min_row_table, max_row_table, asset_group_vul_columns, False, False, False, True, f"{get_column_letter(barchart_start_column_value + round(asset_group_vul_chart.width / 1.7)+1)}38")
    charts_length.append(barchart_start_column_value + round(asset_group_vul_chart.width / 1.7)+1+round(asset_group_vul_piechart.width / 1.7))
    min_row_table = min_row_table + asset_group_vul_rows + 2
    max_row_table = max_row_table + 2 + family_vul_rows
    family_vul_chart = barchart_creation(sheet, "col", "Family", "Family", "Count", 2, family_vul_columns, min_row_table, max_row_table, True, False, False, False, f"{barchart_start_column}55", chartStyle=10, shape=4)
    family_vul_piechart = piechart_creation(sheet, 4, "Family", 2, min_row_table, max_row_table, family_vul_columns, False, False, False, True, f"{get_column_letter(barchart_start_column_value + round(family_vul_chart.width / 1.7)+1)}55")
    charts_length.append(barchart_start_column_value + round(family_vul_chart.width / 1.7)+1+round(family_vul_piechart.width / 1.7))

    merge_cells_title(sheet, "A1", "B2", 1, 1, "Summary Table", "center", "center", "00FFFF00")
    merge_cells_title(sheet, "D1", f"{get_column_letter(max(charts_length))}2", 1, 4, "Summary Table Charts", "center", "center", '0000FF00')
    merge_cells_title(sheet, "A30", "B31", 30, 1, "Click here to see all variables", "center", "center", "FFCCCC")

    wb.save(newfile)

    sheet_name = "Plugin Family"
    book = load_workbook(newfile, keep_vba=True)
    with pd.ExcelWriter(newfile, engine="openpyxl", mode="a", if_sheet_exists="overlay", engine_kwargs={"keep_vba": True}) as writer:
        sheet = sheet_name
        new_row = 3
        pivot_table_1_wide.to_excel(writer, sheet_name=sheet, startrow=new_row, startcol=0, index=False)
        ws = writer.sheets[sheet]

        header_row = new_row+1
        first_data_row = header_row+1
        last_data_row = ws.max_row
        start_col = 1
        last_col = pivot_table_1_wide_columns

        ws.auto_filter.ref = f"{get_column_letter(start_col)}{header_row}:{get_column_letter(last_col)}{last_data_row}"

        wb = writer.book
        ws = wb[sheet]
        autosize_df_columns(ws, pivot_table_1_wide)

    wb = load_workbook(newfile, keep_vba=True)
    sheet = wb[sheet_name]

    min_row_table = 4
    max_row_table = min_row_table + pivot_table_1_wide_rows
    pivot_table_1_wide_chart = barchart_creation(sheet, "col", "Severity and Plugin Family", "Family", "Count", 2, pivot_table_1_wide_columns, min_row_table, max_row_table, True, False, False, False, f"{get_column_letter(pivot_table_1_wide_columns+3)}4", chartGrouping="percentStacked", chartOverlap=100)
    
    merge_cells_title(sheet, "A1", f"{get_column_letter(pivot_table_1_wide_columns)}2", 1, 1, f"{sheet_name} Table", "center", "center", "00FFFF00")
    merge_cells_title(sheet, f"{get_column_letter(pivot_table_1_wide_columns+3)}1", f"{get_column_letter(pivot_table_1_wide_columns+3+round(pivot_table_1_wide_chart.width / 1.7))}2", 1, pivot_table_1_wide_columns+3, "Summary Table Charts", "center", "center", '0000FF00')

    wb.save(newfile)

    sheet_name = "Asset Group"
    book = load_workbook(newfile, keep_vba=True)
    with pd.ExcelWriter(newfile, engine="openpyxl", mode="a", if_sheet_exists="overlay", engine_kwargs={"keep_vba": True}) as writer:
        sheet = sheet_name
        new_row = 3
        pivot_table_2_wide.to_excel(writer, sheet_name=sheet, startrow=new_row, startcol=0, index=False)
        ws = writer.sheets[sheet]

        header_row = new_row+1
        first_data_row = header_row+1
        last_data_row = ws.max_row
        start_col = 1
        last_col = pivot_table_2_wide_columns

        ws.auto_filter.ref = f"{get_column_letter(start_col)}{header_row}:{get_column_letter(last_col)}{last_data_row}"
        hyperlink_cell(ws, first_data_row, last_data_row)

        wb = writer.book
        ws = wb[sheet]
        autosize_df_columns(ws, pivot_table_2_wide)

    wb = load_workbook(newfile, keep_vba=True)
    sheet = wb[sheet_name]

    min_row_table = 4
    max_row_table = min_row_table + pivot_table_2_wide_rows
    pivot_table_2_wide_chart = barchart_creation(sheet, "col", "Severity and Asset Group", "Asset Group", "Count", 2, pivot_table_2_wide_columns, min_row_table, max_row_table, True, False, False, False, f"{get_column_letter(pivot_table_2_wide_columns+3)}4", chartGrouping="percentStacked", chartOverlap=100)
    
    merge_cells_title(sheet, "A1", f"{get_column_letter(pivot_table_2_wide_columns)}2", 1, 1, f"{sheet_name} Table", "center", "center", "00FFFF00")
    merge_cells_title(sheet, f"{get_column_letter(pivot_table_2_wide_columns+3)}1", f"{get_column_letter(pivot_table_2_wide_columns+3+round(pivot_table_2_wide_chart.width / 1.7))}2", 1, pivot_table_2_wide_columns+3, "Summary Table Charts", "center", "center", '0000FF00')

    wb.save(newfile)

    sheet_name = "Overdue"
    book = load_workbook(newfile, keep_vba=True)
    with pd.ExcelWriter(newfile, engine="openpyxl", mode="a", if_sheet_exists="overlay", engine_kwargs={"keep_vba": True}) as writer:
        sheet = sheet_name
        new_row = 3
        pivot_table_3_wide.to_excel(writer, sheet_name=sheet, startrow=new_row, startcol=0, index=False)
        ws = writer.sheets[sheet]

        header_row = new_row+1
        first_data_row = header_row+1
        last_data_row = ws.max_row
        start_col = 1
        last_col = pivot_table_3_wide_columns

        ws.auto_filter.ref = f"{get_column_letter(start_col)}{header_row}:{get_column_letter(last_col)}{last_data_row}"
        hyperlink_cell(ws, first_data_row, last_data_row)

        wb = writer.book
        ws = wb[sheet]
        autosize_df_columns(ws, pivot_table_3_wide)

    wb = load_workbook(newfile, keep_vba=True)
    sheet = wb[sheet_name]

    min_row_table = 4
    max_row_table = min_row_table + pivot_table_3_wide_rows
    pivot_table_3_wide_chart = barchart_creation(sheet, "col", "Overdue and Plugin Family", "Overdue", "Count", 2, pivot_table_3_wide_columns, min_row_table, max_row_table, True, False, False, False, f"{get_column_letter(pivot_table_3_wide_columns+3)}4", chartGrouping="percentStacked", chartOverlap=100)
    
    merge_cells_title(sheet, "A1", f"{get_column_letter(pivot_table_3_wide_columns)}2", 1, 1, f"{sheet_name} Table", "center", "center", "00FFFF00")
    merge_cells_title(sheet, f"{get_column_letter(pivot_table_3_wide_columns+3)}1", f"{get_column_letter(pivot_table_3_wide_columns+3+round(pivot_table_3_wide_chart.width / 1.7))}2", 1, pivot_table_3_wide_columns+3, "Summary Table Charts", "center", "center", '0000FF00')

    wb.save(newfile)

    sheet_name = "Asset Grp & Family"
    book = load_workbook(newfile, keep_vba=True)
    with pd.ExcelWriter(newfile, engine="openpyxl", mode="a", if_sheet_exists="overlay", engine_kwargs={"keep_vba": True}) as writer:
        sheet = sheet_name
        new_row = 3
        pivot_table_4_wide.to_excel(writer, sheet_name=sheet, startrow=new_row, startcol=0, index=False)
        ws = writer.sheets[sheet]

        header_row = new_row+1
        first_data_row = header_row+1
        last_data_row = ws.max_row
        start_col = 1
        last_col = pivot_table_4_wide_columns

        ws.auto_filter.ref = f"{get_column_letter(start_col)}{header_row}:{get_column_letter(last_col)}{last_data_row}"
        hyperlink_cell(ws, first_data_row, last_data_row, column_no=3)

        wb = writer.book
        ws = wb[sheet]
        autosize_df_columns(ws, pivot_table_4_wide)

    wb = load_workbook(newfile, keep_vba=True)
    sheet = wb[sheet_name]

    min_row_table = 4
    max_row_table = min_row_table + pivot_table_4_wide_rows
    pivot_table_4_wide_chart = barchart_creation(sheet, "col", "Asset Group, Plugin Family and Severity", "Asset Group - Plugin Family", "Count", 4, pivot_table_4_wide_columns, min_row_table, max_row_table, True, False, False, False, f"{get_column_letter(pivot_table_4_wide_columns+3)}4", chartGrouping="percentStacked", chartOverlap=100)
    
    merge_cells_title(sheet, "A1", f"{get_column_letter(pivot_table_4_wide_columns)}2", 1, 1, f"{sheet_name} Table", "center", "center", "00FFFF00")
    merge_cells_title(sheet, f"{get_column_letter(pivot_table_4_wide_columns+3)}1", f"{get_column_letter(pivot_table_4_wide_columns+3+round(pivot_table_4_wide_chart.width / 1.7))}2", 1, pivot_table_4_wide_columns+3, "Summary Table Charts", "center", "center", '0000FF00')

    wb.save(newfile)

if __name__ == "__main__":
    main()